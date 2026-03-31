from __future__ import annotations

import csv
import io
import os
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError, transaction
from django.db.utils import OperationalError
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.http import require_GET, require_POST
from django.db.models import Q

from accounts.roles import can_delete as erp_can_delete, can_edit as erp_can_edit
from commerce.models import Product
from khataapp.models import Party
from khataapp.utils.whatsapp_utils import send_whatsapp_message
from procurement.forms import SupplierForm, SupplierPriceUploadForm, SupplierProductForm
from procurement.models import SupplierPriceAlert, SupplierPriceHistory, SupplierProduct, SupplierRating
from procurement.services import best_supplier_for_product, best_supplier_map_for_products, rank_suppliers_for_product, supplier_ratings_map


def _supplier_qs_for_user(user):
    return Party.objects.filter(owner=user, party_type="supplier").order_by("name", "id")

def _missing_tables_message() -> str:
    return "Procurement module database tables are missing. Please run migrations: manage.py migrate procurement"


@login_required
@require_GET
def supplier_list(request):
    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip().lower()

    qs = _supplier_qs_for_user(request.user)
    if q:
        qs = qs.filter(name__icontains=q)
    if status == "active":
        qs = qs.filter(is_active=True)
    elif status == "inactive":
        qs = qs.filter(is_active=False)

    suppliers = list(qs)
    try:
        rating_map = supplier_ratings_map(request.user, [int(s.id) for s in suppliers])
    except OperationalError:
        rating_map = {}
        messages.error(request, _missing_tables_message())
    for s in suppliers:
        info = rating_map.get(int(s.id), {}) if rating_map else {}
        setattr(s, "avg_rating", info.get("avg") or Decimal("0.00"))
        setattr(s, "rating_count", int(info.get("count") or 0))

    context = {
        "suppliers": suppliers,
        "q": q,
        "status": status,
    }
    return render(request, "procurement/supplier_list.html", context)


@login_required
def supplier_add(request):
    if not erp_can_edit(request.user):
        messages.error(request, "Permission denied: view-only role cannot add suppliers.")
        return redirect("procurement:supplier_list")

    form = SupplierForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            supplier: Party = form.save(commit=False)
            supplier.owner = request.user
            supplier.party_type = "supplier"
            supplier.save()
            messages.success(request, "Supplier created.")
            return redirect("procurement:supplier_list")

    return render(request, "procurement/supplier_add.html", {"form": form})


@login_required
def supplier_edit(request, supplier_id: int):
    supplier = get_object_or_404(Party, id=supplier_id, owner=request.user, party_type="supplier")

    if not erp_can_edit(request.user):
        messages.error(request, "Permission denied: view-only role cannot edit suppliers.")
        return redirect("procurement:supplier_list")

    form = SupplierForm(request.POST or None, instance=supplier)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Supplier updated.")
            return redirect("procurement:supplier_list")

    return render(request, "procurement/supplier_edit.html", {"form": form, "supplier": supplier})


@login_required
@require_POST
def supplier_delete(request, supplier_id: int):
    supplier = get_object_or_404(Party, id=supplier_id, owner=request.user, party_type="supplier")

    if not erp_can_delete(request.user):
        messages.error(request, "Permission denied: only Admin can delete.")
        return redirect("procurement:supplier_list")

    # Safety: never hard-delete suppliers (would cascade to orders/ledger). Deactivate instead.
    supplier.is_active = False
    supplier.save(update_fields=["is_active"])
    messages.success(request, "Supplier deactivated.")
    return redirect("procurement:supplier_list")


@login_required
def supplier_product_mapping(request):
    if not erp_can_edit(request.user):
        messages.error(request, "Permission denied: view-only role cannot manage supplier mapping.")
        return redirect("accounts:dashboard")

    q = (request.GET.get("q") or "").strip()
    supplier_id = (request.GET.get("supplier_id") or "").strip()
    product_id = (request.GET.get("product_id") or "").strip()
    mapping_id = (request.GET.get("mapping_id") or "").strip()

    mapping_obj = None
    if mapping_id:
        try:
            mapping_obj = SupplierProduct.objects.select_related("supplier", "product").get(id=int(mapping_id), owner=request.user)
        except Exception:
            mapping_obj = None

    form = SupplierProductForm(request.POST or None, instance=mapping_obj, owner=request.user)
    upload_form = SupplierPriceUploadForm(request.POST or None, request.FILES or None)

    if request.method == "POST" and request.POST.get("action") == "save_mapping":
        if form.is_valid():
            obj: SupplierProduct = form.save(commit=False)
            obj.owner = request.user
            obj._updated_by = request.user
            try:
                obj.save()
            except OperationalError:
                messages.error(request, _missing_tables_message())
                return redirect("procurement:supplier_product_mapping")
            except IntegrityError:
                # Unique mapping exists -> update it instead
                try:
                    existing = SupplierProduct.objects.filter(
                        owner=request.user, supplier=obj.supplier, product=obj.product
                    ).first()
                except OperationalError:
                    messages.error(request, _missing_tables_message())
                    return redirect("procurement:supplier_product_mapping")
                if existing:
                    existing.price = obj.price
                    existing.moq = obj.moq
                    existing.delivery_days = obj.delivery_days
                    existing.is_active = obj.is_active
                    existing._updated_by = request.user
                    try:
                        existing.save()
                    except OperationalError:
                        messages.error(request, _missing_tables_message())
                        return redirect("procurement:supplier_product_mapping")
                else:
                    raise
            messages.success(request, "Supplier product mapping saved.")
            return redirect("procurement:supplier_product_mapping")
        messages.error(request, "Please fix the errors in the form.")

    if request.method == "POST" and request.POST.get("action") == "upload_prices":
        if upload_form.is_valid():
            f = upload_form.cleaned_data["file"]
            try:
                created, updated, skipped = _import_price_list(request.user, f, updated_by=request.user)
                messages.success(request, f"Price list imported. Created: {created}, Updated: {updated}, Skipped: {skipped}.")
            except OperationalError:
                messages.error(request, _missing_tables_message())
            except Exception as e:
                messages.error(request, f"Upload failed: {type(e).__name__}: {e}")
            return redirect("procurement:supplier_product_mapping")
        messages.error(request, "Please choose a valid file.")

    try:
        mappings = SupplierProduct.objects.filter(owner=request.user).select_related("supplier", "product")
        if q:
            mappings = mappings.filter(
                Q(product__name__icontains=q) | Q(product__sku__icontains=q) | Q(supplier__name__icontains=q)
            )
        if supplier_id:
            try:
                mappings = mappings.filter(supplier_id=int(supplier_id))
            except Exception:
                pass
        if product_id:
            try:
                mappings = mappings.filter(product_id=int(product_id))
            except Exception:
                pass
        mappings = mappings.order_by("-last_updated", "-id")[:500]
    except OperationalError:
        messages.error(request, _missing_tables_message())
        mappings = SupplierProduct.objects.none()

    context = {
        "form": form,
        "upload_form": upload_form,
        "mappings": mappings,
        "suppliers": _supplier_qs_for_user(request.user),
        "products": Product.objects.filter(owner=request.user).order_by("name", "id"),
        "q": q,
        "supplier_id": supplier_id,
        "product_id": product_id,
        "mapping_obj": mapping_obj,
    }
    return render(request, "procurement/supplier_product_mapping.html", context)


@login_required
@require_GET
def supplier_price_comparison(request):
    q = (request.GET.get("q") or "").strip()
    product_id = (request.GET.get("product_id") or "").strip()

    product = None
    if product_id:
        try:
            product = Product.objects.filter(owner=request.user, id=int(product_id)).first()
        except Exception:
            product = None

    products = Product.objects.filter(owner=request.user).order_by("name", "id")
    if q:
        products = products.filter(Q(name__icontains=q) | Q(sku__icontains=q))
    products = products[:30]

    ranked = []
    best = None
    alerts = []
    if product:
        try:
            ranked = rank_suppliers_for_product(owner=request.user, product_id=product.id)
            best = ranked[0] if ranked else None
            alerts = list(
                SupplierPriceAlert.objects.filter(owner=request.user, product=product)
                .select_related("supplier")
                .order_by("-created_at", "-id")[:10]
            )
        except OperationalError:
            messages.error(request, _missing_tables_message())
            ranked = []
            best = None
            alerts = []

    context = {
        "q": q,
        "products": products,
        "product": product,
        "ranked": ranked,
        "best": best,
        "alerts": alerts,
    }
    return render(request, "procurement/supplier_price_comparison.html", context)


@login_required
@require_POST
def mark_alert_read(request, alert_id: int):
    try:
        alert = get_object_or_404(SupplierPriceAlert, id=alert_id, owner=request.user)
        alert.is_read = True
        alert.save(update_fields=["is_read"])
    except OperationalError:
        messages.error(request, _missing_tables_message())
    return redirect(request.META.get("HTTP_REFERER") or reverse("procurement:supplier_price_comparison"))


@login_required
@require_GET
def api_low_stock_best_suppliers(request):
    """
    Lightweight JSON helper for low-stock screens.
    """
    product_ids_raw = (request.GET.get("product_ids") or "").strip()
    try:
        product_ids = [int(x) for x in product_ids_raw.split(",") if x.strip().isdigit()]
    except Exception:
        product_ids = []

    try:
        best_map = best_supplier_map_for_products(request.user, product_ids)
    except OperationalError:
        return JsonResponse({"detail": _missing_tables_message()}, status=503)
    payload = {
        str(pid): {
            "supplier_id": best_map[pid].supplier_id,
            "supplier_name": best_map[pid].supplier_name,
            "price": str(best_map[pid].price),
            "moq": best_map[pid].moq,
            "delivery_days": best_map[pid].delivery_days,
        }
        for pid in best_map
    }
    return JsonResponse(payload)


@login_required
@require_POST
def api_send_whatsapp_order(request):
    supplier_id_raw = (request.POST.get("supplier_id") or "").strip()
    product_id_raw = (request.POST.get("product_id") or "").strip()
    qty_raw = (request.POST.get("qty") or "").strip()

    try:
        supplier_id = int(supplier_id_raw)
        product_id = int(product_id_raw)
        qty = int(qty_raw or "1")
    except Exception:
        return JsonResponse({"ok": False, "detail": "Invalid input."}, status=400)

    supplier = get_object_or_404(Party, id=supplier_id, owner=request.user, party_type="supplier")
    product = get_object_or_404(Product, id=product_id, owner=request.user)

    try:
        best = best_supplier_for_product(owner=request.user, product_id=product.id)
    except OperationalError:
        best = None
    price = best.price if best and best.supplier_id == supplier.id else Decimal(str(request.POST.get("price") or "0"))

    to_number = (supplier.whatsapp_number or supplier.mobile or "").strip()
    if not to_number:
        return JsonResponse({"ok": False, "detail": "Supplier WhatsApp/Mobile number not set."}, status=400)

    total = (Decimal(str(qty)) * Decimal(str(price or 0))).quantize(Decimal("0.01"))
    message = (
        "Purchase Order Request\n"
        f"Supplier: {supplier.name}\n"
        f"Product: {product.name} ({product.sku})\n"
        f"Qty: {qty}\n"
        f"Rate: ₹{price}\n"
        f"Total: ₹{total}\n"
        "\nReply with availability & expected delivery date."
    )

    status_code, resp_text = send_whatsapp_message(to_number, message)
    return JsonResponse({"ok": int(status_code) in {200, 201, 202}, "status_code": status_code, "response": resp_text})


def _import_price_list(owner, uploaded_file, updated_by=None) -> tuple[int, int, int]:
    """
    Returns: (created, updated, skipped)
    """
    name = str(getattr(uploaded_file, "name", "") or "")
    ext = os.path.splitext(name.lower())[1]

    rows = []
    if ext in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except Exception as e:
            raise RuntimeError("openpyxl is required for Excel uploads") from e

        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.active
        headers = []
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx == 0:
                headers = [str(x or "").strip().lower() for x in row]
                continue
            if not any(row):
                continue
            d = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            rows.append(d)
    else:
        # CSV (default)
        content = uploaded_file.read()
        try:
            text = content.decode("utf-8-sig")
        except Exception:
            text = content.decode("latin-1")
        reader = csv.DictReader(io.StringIO(text))
        for r in reader:
            rows.append({(k or "").strip().lower(): (v or "").strip() for k, v in r.items()})

    created = 0
    updated = 0
    skipped = 0

    with transaction.atomic():
        for r in rows:
            supplier_name = str(r.get("supplier") or r.get("supplier_name") or "").strip()
            sku = str(r.get("product_sku") or r.get("sku") or "").strip()
            product_name = str(r.get("product_name") or r.get("product") or "").strip()
            price_raw = r.get("price") or r.get("purchase_price") or ""
            moq_raw = r.get("moq") or r.get("minimum_order_qty") or r.get("minimum_order_quantity") or "1"
            delivery_raw = r.get("delivery_days") or r.get("delivery_time") or r.get("delivery") or "0"

            if not supplier_name:
                skipped += 1
                continue

            supplier, _ = Party.objects.get_or_create(
                owner=owner,
                party_type="supplier",
                name=supplier_name,
                defaults={"is_active": True},
            )

            product = None
            if sku:
                product = Product.objects.filter(owner=owner, sku=sku).first()
            if not product and product_name:
                product = Product.objects.filter(owner=owner, name__iexact=product_name).first()
            if not product:
                skipped += 1
                continue

            try:
                price = Decimal(str(price_raw or "0")).quantize(Decimal("0.01"))
            except Exception:
                price = Decimal("0.00")

            try:
                moq = int(str(moq_raw or "1"))
            except Exception:
                moq = 1
            try:
                delivery_days = int(str(delivery_raw or "0"))
            except Exception:
                delivery_days = 0

            mapping = SupplierProduct.objects.filter(owner=owner, supplier=supplier, product=product).first()
            if not mapping:
                mapping = SupplierProduct(owner=owner, supplier=supplier, product=product)
                created += 1
            else:
                updated += 1

            mapping.price = price
            mapping.moq = max(moq, 1)
            mapping.delivery_days = max(delivery_days, 0)
            mapping.is_active = True
            mapping._updated_by = updated_by
            mapping.save()

    return created, updated, skipped
